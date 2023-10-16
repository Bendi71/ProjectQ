from bs4 import BeautifulSoup
import requests
import time
from openpyxl import load_workbook
from datetime import datetime

#inputok
Tickers=[]
idokoz=int(input('Rerun scraping in minutes: '))
n = int(input("Number of tickers: "))
for i in range(0, n):
    elem = input(f'ticker no. {i}: ')
    Tickers.append(elem)

if __name__ == '__main__':
    while True:
        for ticker in Tickers:
            if ticker[0] == "^": #ez az indexekhez kell
                ticker = ticker[1:]
                html_text = requests.get(f"https://finance.yahoo.com/quote/%5E{ticker}?p=^{ticker}").text
            else:
                html_text = requests.get(f"https://finance.yahoo.com/quote/{ticker}?p={ticker}").text
            soup = BeautifulSoup(html_text, 'lxml')

            # ár
            pricecard = soup.find('div', class_='D(ib) Va(m) Maw(65%) Ov(h)') #az árat is csak csúsztatva tudja lehúzni
            price = pricecard.find('fin-streamer', class_='Fw(b) Fz(36px) Mb(-4px) D(ib)').text

            # nyitva van-e a piac
            datecard = pricecard.find('div', class_='C($tertiaryColor) D(b) Fz(12px) Fw(n) Mstart(0)--mobpsm Mt(6px)--mobpsm Whs(n)').text

            # idő
            date = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

            # adattáblából egyéb adatok, ezt jelenleg nem írja ki
            cardbody = soup.find('table', class_='W(100%)')
            cardall = cardbody.find_all('tr', class_='Bxz(bb) Bdbw(1px) Bdbs(s) Bdc($seperatorColor) H(36px)')
            data = []
            for index in cardall:
                data.append(index.find('td', class_='Ta(end) Fw(600) Lh(14px)').text)
                # volument csak késleltetve tudja lehúzni

            # itt kezdődik az excelezés
            # mindegyik részvény adata külön munkalapra kerül
            wb = load_workbook('yahoofinance.xlsx')
            if ticker not in wb.sheetnames:
                wb.create_sheet(ticker)
                wb.active = wb[ticker]
                ws = wb.active
                ws['A1'] = "Date"
                ws['B1'] = "Price"
                if "At close" not in datecard:
                    lepes = ws.max_row
                    ws[f'A{lepes + 1}'] = date
                    ws[f'B{lepes + 1}'] = price
                else:
                    print("Market is closed")
            else:
                wb.active = wb[ticker]
                ws = wb.active
                if "At close" not in datecard:
                    lepes = ws.max_row
                    ws[f'A{lepes + 1}'] = date
                    ws[f'B{lepes + 1}'] = price
                else:
                    print("Market is closed")
            wb.save('yahoofinance.xlsx')
        time_wait=idokoz
        print(f'Waiting {time_wait} minutes...')
        time.sleep(time_wait*60)